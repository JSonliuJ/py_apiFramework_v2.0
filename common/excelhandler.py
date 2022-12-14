# -- encoding: utf-8 --
# @time:    	2022/4/6 22:24
# @Author: 		JsonLiu
# @Email:  		492224300@qq.com
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import openpyxl

class ExcelHandle:
    """

    """
    def __init__(self,file_name):
     # 1.打开工作簿
        self.wb = openpyxl.load_workbook(file_name)
    def _open_sheet(self,sheet_name) -> Worksheet:
        # 2.打开表单
        '''
        在函数或方法后面加 -> 类型: 表示此函数返回值是一个 xxx的类型
        '''
        sheet = self.wb[sheet_name]
        self.wb.close()
        return sheet

    def read_tab_header(self,sheet_name):
        # 3.读取标题
        sheet = self._open_sheet(sheet_name)
        # tab_header = []
        # for i in sheet[1]:
        #     tab_header.append(i.value)
        tab_header = list(sheet.values)[0]
        return tab_header
    def read_all_data(self,sheet_name):
        # 4.读取所有数据：方式1
        # 列表转字典：和header 进行zip
        global dict_data
        sheet = self._open_sheet(sheet_name)
        rows = list(sheet.rows)
        all_data = []
        # 标题 header = sheet.rows[0]
        for row in rows[1:]:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
                dict_data = dict(zip(self.read_tab_header(sheet_name),row_data))
            all_data.append(dict_data)
        return all_data

    def read_all_data_lc(self,sheet_name):
        # 列表推导式
        sheet = self._open_sheet(sheet_name)
        rows = list(sheet.values)
        tab_head = rows[0]
        all_data = [dict(zip(tab_head,i)) for i in rows[1:]]
        return all_data
    @staticmethod
    def write_back(file_name,sheet_name,row,column,write_value):
        '''
        # 5. 指定单元格写入(使用静态方法，不要使用实例方法)
        '''
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        # 修改单元格
        sheet.cell(row,column).value = write_value
        # 保存
        wb.save(file_name)
        # 关闭
        wb.close()

if __name__ == '__main__':
    from common.filehandle import test_data_path
    EH = ExcelHandle(file_name=test_data_path)
    # all_data = EH.read_all_data('register')
    # print(all_data)
    data = EH.read_all_data_lc('register')
    print(data)
